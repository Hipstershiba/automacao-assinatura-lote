from datetime import datetime
import os
import logging
import unicodedata

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False


def _sanear_nome(usuario_nome):
    if not usuario_nome:
        return 'usuario'

    nome_norm = unicodedata.normalize('NFKD', usuario_nome.strip())
    valor = ''.join(
        c if c.isalnum() else '_'
        for c in nome_norm
        if not unicodedata.combining(c)
    )
    return valor or 'usuario'


class GeradorRelatorio:
    def __init__(self, usuario_cpf, usuario_nome, filtros, ordem_assinatura):
        """Inicializa o gerador de relatório."""
        self.usuario_cpf = usuario_cpf
        self.usuario_nome = _sanear_nome(usuario_nome)
        self.filtros = filtros or {}
        self.ordem_assinatura = ordem_assinatura or ''
        self.data_geracao = datetime.now()
        self.contratos_assinados = []
        self.contratos_pulados = []
        self.sessao_inicio = self.data_geracao.strftime('%d/%m/%Y %H:%M:%S')
        self.sessao_data = self.data_geracao.strftime('%Y-%m-%d')
        self.sessao_hora = self.data_geracao.strftime('%H:%M:%S')

    def registrar_assinatura(self, contrato, horario_assinatura):
        """Registra um contrato assinado com sucesso."""
        self.contratos_assinados.append({
            'titulo': contrato.titulo,
            'codigo': contrato.codigo,
            'data_criacao': contrato.data_criacao,
            'data_limite': contrato.data_limite,
            'horario_assinatura': horario_assinatura
        })

    def registrar_pulado(self, contrato, motivo):
        """Registra um contrato que foi pulado."""
        self.contratos_pulados.append({
            'titulo': contrato.titulo,
            'codigo': contrato.codigo,
            'data_criacao': contrato.data_criacao,
            'data_limite': contrato.data_limite,
            'motivo': motivo,
            'horario': datetime.now().strftime('%H:%M:%S')
        })

    def salvar_relatorio(self, diretorio='relatorios'):
        """Salva o relatório em Excel por sessão."""
        if not OPENPYXL_DISPONIVEL:
            raise RuntimeError('Biblioteca openpyxl não está disponível; instale openpyxl para gerar relatórios Excel.')

        os.makedirs(diretorio, exist_ok=True)
        data_formatada = self.data_geracao.strftime('%Y-%m-%d_%Hh%Mm%Ss')
        return self._salvar_excel(diretorio, data_formatada)

    def _salvar_excel(self, diretorio, data_formatada):
        nome_arquivo = os.path.join(diretorio, f'relatorio_{self.usuario_nome}_{data_formatada}.xlsx')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._criar_aba_resumo(wb)
        self._criar_aba_assinados(wb)
        self._criar_aba_pulados(wb)

        self._append_assinados(wb['Assinados'])
        self._append_pulados(wb['Pulados'])
        self._append_resumo(wb['Resumo'], wb)
        wb.save(nome_arquivo)
        logging.info(f'Relatório Excel salvo em: {nome_arquivo}')
        print(f'[INFO] Relatório Excel salvo em: {nome_arquivo}')
        return nome_arquivo

    def _criar_aba_resumo(self, wb):
        if 'Resumo' in wb.sheetnames:
            wb.remove(wb['Resumo'])

        ws = wb.create_sheet('Resumo')
        titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        titulo_font = Font(bold=True, color='FFFFFF', size=12)
        cabecalho_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cabecalho_font = Font(bold=True, size=11)

        ws['A1'] = 'RELATÓRIO DIÁRIO DE ASSINATURA EM LOTE'
        ws['A1'].font = titulo_font
        ws['A1'].fill = titulo_fill
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws['A2'] = f'Data do relatório:'
        ws['B2'] = self.sessao_data
        ws['A3'] = f'Gerado em:'
        ws['B3'] = self.sessao_hora
        ws['A4'] = 'Assinante (Nome):'
        ws['B4'] = self.usuario_nome
        ws['A5'] = 'Ordem de Assinatura:'
        ws['B5'] = self.ordem_assinatura

        ws['A7'] = 'FILTROS UTILIZADOS'
        ws['A7'].font = cabecalho_font
        ws['A7'].fill = cabecalho_fill
        row = 8
        if self.filtros.get('data_inicial'):
            ws[f'A{row}'] = 'Data Inicial:'
            ws[f'B{row}'] = self.filtros['data_inicial']
            row += 1
        if self.filtros.get('data_final'):
            ws[f'A{row}'] = 'Data Final:'
            ws[f'B{row}'] = self.filtros['data_final']
            row += 1
        if self.filtros.get('tag'):
            ws[f'A{row}'] = 'Tag:'
            ws[f'B{row}'] = self.filtros['tag']
            row += 1
        if self.filtros.get('nome_documento'):
            ws[f'A{row}'] = 'Nome do Documento:'
            ws[f'B{row}'] = self.filtros['nome_documento']
            row += 1

        row += 1
        ws[f'A{row}'] = 'RESUMO'
        ws[f'A{row}'].font = cabecalho_font
        ws[f'A{row}'].fill = cabecalho_fill
        row += 1
        ws[f'A{row}'] = 'Total Processados:'
        ws[f'B{row}'] = 0
        row += 1
        ws[f'A{row}'] = 'Assinados com Sucesso:'
        ws[f'B{row}'] = 0
        row += 1
        ws[f'A{row}'] = 'Pulados:'
        ws[f'B{row}'] = 0

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        return ws

    def _criar_aba_assinados(self, wb):
        ws = wb.create_sheet('Assinados')
        cabecalho_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        cabecalho_font = Font(bold=True, color='000000')
        headers = ['Título', 'Código', 'Data de Criação', 'Data Limite', 'Horário de Assinatura']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        return ws

    def _criar_aba_pulados(self, wb):
        ws = wb.create_sheet('Pulados')
        cabecalho_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        cabecalho_font = Font(bold=True, color='000000')
        headers = ['Título', 'Código', 'Data de Criação', 'Data Limite', 'Motivo', 'Horário']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
        return ws

    def _append_assinados(self, ws):
        for contrato in self.contratos_assinados:
            ws.append([
                contrato['titulo'],
                contrato['codigo'],
                contrato['data_criacao'],
                contrato['data_limite'],
                contrato['horario_assinatura']
            ])

    def _append_pulados(self, ws):
        for contrato in self.contratos_pulados:
            ws.append([
                contrato['titulo'],
                contrato['codigo'],
                contrato['data_criacao'],
                contrato['data_limite'],
                contrato['motivo'],
                contrato['horario']
            ])

    def _append_resumo(self, ws, wb):
        self._atualizar_totais(ws)

    def _atualizar_totais(self, ws):
        total_assinados = self._contar_linhas(ws.parent, 'Assinados')
        total_pulados = self._contar_linhas(ws.parent, 'Pulados')
        total_processados = total_assinados + total_pulados
        self._definir_valor_por_rotulo(ws, 'Total Processados:', total_processados)
        self._definir_valor_por_rotulo(ws, 'Assinados com Sucesso:', total_assinados)
        self._definir_valor_por_rotulo(ws, 'Pulados:', total_pulados)

    def _definir_valor_por_rotulo(self, ws, rotulo, valor):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and str(cell.value).strip() == rotulo:
                ws.cell(row=row, column=2).value = valor
                return

        # fallback para o caso inesperado de a estrutura do resumo variar
        if rotulo == 'Total Processados:':
            ws['B13'] = valor
        elif rotulo == 'Assinados com Sucesso:':
            ws['B14'] = valor
        elif rotulo == 'Pulados:':
            ws['B15'] = valor

    def _contar_linhas(self, wb, nome_aba):
        if nome_aba not in wb.sheetnames:
            return 0
        ws = wb[nome_aba]
        return max(ws.max_row - 1, 0)

